import os
import sys
import re
import json
import time
import random
import sqlite3
import requests
from datetime import date
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from dotenv import load_dotenv

load_dotenv()

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import (
    DB_PATH, TARGET_ROLES,
    AUTO_APPLY_THRESHOLD, DREAM_JOB_THRESHOLD, MAYBE_THRESHOLD,
    NCR_CITIES, DREAM_COMPANIES, MIN_SALARY_LPA,
    DREAM_MANUAL_DIR, SALARY_MANUAL_DIR,
    NCR_SEARCH_LOCATIONS, REMOTE_SEARCH_LOCATIONS,
    OUTPUTS_DIR, DEV_MAX_JOBS_PER_RUN, HEADLESS_MODE, MISSION_JOB_BATCH_SIZE, GENERATE_DOCUMENTS_DEFAULT
)
from job_scorer import score_job
from resume_tailor import generate_manual_package, tailor_resume
from database.db import insert_job, update_job_with_documents

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
}

def extract_job_external_id(job_url):
    """Extract stable job ID from LinkedIn URL"""
    if not job_url:
        return None
    
    # Pattern 1: /jobs/view/<id>/
    match = re.search(r'/jobs/view/(\d+)', job_url)
    if match:
        return match.group(1)
    
    # Pattern 2: currentJobId=<id> parameter
    match = re.search(r'currentJobId=(\d+)', job_url)
    if match:
        return match.group(1)
    
    # Pattern 3: /jobs/view/<id>/...
    match = re.search(r'/jobs/view/(\d+)', job_url)
    if match:
        return match.group(1)
    
    # Pattern 4: General numeric ID in URL path
    match = re.search(r'(\d{6,})', job_url)  # LinkedIn job IDs are usually 6+ digits
    if match:
        return match.group(1)
    
    print(f"  ⚠️ Could not extract job_external_id from: {job_url}")
    return None

def is_remote_location(location):
    """Check if location indicates remote work"""
    if not location:
        return False
    return "remote" in location.lower() or "anywhere" in location.lower()

def extract_apply_by_date(jd_text):
    """Extract apply by deadline from job description"""
    if not jd_text:
        return None
    
    # Common patterns for application deadlines
    patterns = [
        r'apply by[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'application deadline[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'apply before[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'expires[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'closing date[:\s]*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, jd_text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
    
    return None

def estimate_expected_salary(salary_range):
    """Extract expected salary from salary range text"""
    if not salary_range:
        return None
    
    # Try to extract numeric range (e.g., "15-25 LPA", "$80k-120k")
    salary_range = salary_range.strip()
    
    # Pattern for LPA format
    lpa_match = re.search(r'(\d+)[\s-]*(\d*)\s*LPA', salary_range, re.IGNORECASE)
    if lpa_match:
        min_salary = int(lpa_match.group(1))
        max_salary = int(lpa_match.group(2)) if lpa_match.group(2) else min_salary
        return f"{min_salary}-{max_salary} LPA" if min_salary != max_salary else f"{min_salary} LPA"
    
    # Pattern for dollar format
    dollar_match = re.search(r'\$?(\d+)[kK]?[\s-]*\$?(\d*)[kK]?', salary_range)
    if dollar_match:
        min_salary = int(dollar_match.group(1))
        max_salary = int(dollar_match.group(2)) if dollar_match.group(2) else min_salary
        return f"${min_salary}k-${max_salary}k" if min_salary != max_salary else f"${min_salary}k"
    
    # Return original if no pattern matched
    return salary_range

EA_SEARCH_LOCATIONS = [
    "Noida, Uttar Pradesh, India",
    "Delhi, India",
    "Gurugram, Haryana, India",
    "Remote"
]


# ─────────────────────────────────────────────
# DRIVER — dedicated chrome_profile folder
# ─────────────────────────────────────────────

def create_driver():
    options = Options()
    options.add_argument(r"--user-data-dir=C:\Users\HP\AutoApply\chrome_profile")
    options.add_argument("--profile-directory=Default")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    # Add headless mode for development with more robust options
    if HEADLESS_MODE:
        options.add_argument("--headless=new")  # Use new headless mode
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-plugins")
        options.add_argument("--disable-images")
        options.add_argument("--disable-javascript")  # Faster for scraping
        print("  🖥️ Running in headless mode")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


# ─────────────────────────────────────────────
# LOGIN — smart: skips if already logged in
# ─────────────────────────────────────────────

def login_linkedin(driver):
    driver.get("https://www.linkedin.com")
    time.sleep(3)

    if "feed" in driver.current_url or "jobs" in driver.current_url or "linkedin.com/in/" in driver.current_url:
        print("  ✅ Already logged into LinkedIn!")
        return True

    print("  🔐 First time login — logging into LinkedIn...")
    driver.get("https://www.linkedin.com/login")
    time.sleep(2)
    try:
        email_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "username"))
        )
        email_field.send_keys(os.environ.get("LINKEDIN_EMAIL", ""))
        driver.find_element(By.ID, "password").send_keys(os.environ.get("LINKEDIN_PASSWORD", ""))
        driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()
        time.sleep(5)

        if "feed" in driver.current_url:
            print("  ✅ Login successful — session saved to chrome_profile!")
            return True
        elif "checkpoint" in driver.current_url:
            print("  ⚠️ LinkedIn checkpoint detected!")
            print("  👉 Complete verification manually in the Chrome window, then press Enter here...")
            input("     Press Enter once you've verified...")
            return True
        else:
            print(f"  ⚠️ Unexpected URL after login: {driver.current_url}")
            return False
    except Exception as e:
        print(f"  ❌ Login error: {e}")
        return False


# ─────────────────────────────────────────────
# FILTERS
# ─────────────────────────────────────────────

def is_ncr_or_remote(location: str) -> bool:
    loc = location.lower()
    if any(word in loc for word in ["remote", "hybrid", "wfh", "work from home"]):
        return True
    if any(city in loc for city in NCR_CITIES):
        return True
    return False


def is_dream_company(company_name: str) -> bool:
    return any(d in company_name.lower() for d in DREAM_COMPANIES)


def extract_salary_lpa(text: str):
    if not text:
        return None
    patterns = [
        r'(\d+)\s*[-–to]+\s*(\d+)\s*(?:lpa|lakh|l\.p\.a)',
        r'(\d+)\s*(?:lpa|lakh|l\.p\.a)',
        r'(?:₹|inr|rs\.?)\s*(\d+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, text.lower())
        if match:
            nums = [int(x) for x in match.groups() if x]
            salary = max(nums)
            if salary < 100:
                return float(salary)
            elif salary >= 100000:
                return round(salary / 100000, 1)
    return None


# ─────────────────────────────────────────────
# SELENIUM HELPERS
# ─────────────────────────────────────────────

def is_easy_apply_page(driver) -> bool:
    try:
        buttons = driver.find_elements(
            By.CSS_SELECTOR,
            "button.jobs-apply-button, .jobs-apply-button--top-card"
        )
        for btn in buttons:
            if "easy apply" in btn.text.lower():
                return True
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            if "easy apply" in btn.text.lower():
                return True
        return False
    except:
        return False


def get_jd_text(driver) -> str:
    try:
        try:
            show_more = driver.find_element(
                By.CSS_SELECTOR, "button.show-more-less-html__button"
            )
            driver.execute_script("arguments[0].click();", show_more)
            time.sleep(1)
        except:
            pass
        selectors = [
            "div.show-more-less-html__markup",
            "div.description__text",
            "div.jobs-description__content",
            "div.jobs-description-content__text",
            "div.jobs-box__html-content"
        ]
        for selector in selectors:
            try:
                elem = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                )
                text = elem.text.strip()
                if len(text) > 200:
                    return text[:5000]
            except:
                continue
        try:
            return driver.find_element(By.TAG_NAME, "main").text.strip()[:5000]
        except:
            return ""
    except:
        return ""


# ─────────────────────────────────────────────
# DATABASE
# ─────────────────────────────────────────────

def save_easy_apply_to_db(job, score_result, resume_path, cover_letter_path):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT app_id FROM applications WHERE company=? AND job_title=?",
        (job["company"], job["title"])
    )
    if cursor.fetchone():
        conn.close()
        return False
    cursor.execute("""
        INSERT INTO applications
        (job_title, company, platform, jd_text, match_score, salary_range,
         applied_date, status, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        job["title"], job["company"], job["platform"],
        job["jd_text"], score_result["score"], job.get("salary_range", ""),
        str(date.today()), "Discovered",
        f'Queue: {score_result["queue"]} | URL: {job.get("url", "")} | '
        f'Resume: {resume_path} | CoverLetter: {cover_letter_path}'
    ))
    conn.commit()
    conn.close()
    return True


def save_manual_to_db(job, queue_type, resume_path, cover_letter_path):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT app_id FROM applications WHERE company=? AND job_title=?",
        (job["company"], job["title"])
    )
    if cursor.fetchone():
        conn.close()
        return False
    cursor.execute("""
        INSERT INTO applications
        (job_title, company, platform, jd_text, match_score, salary_range,
         applied_date, status, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        job["title"], job["company"], job["platform"],
        job["jd_text"], 0, job.get("salary_range", ""),
        str(date.today()), "Manual Pending",
        f'Queue: {queue_type} | URL: {job.get("url", "")} | '
        f'Resume: {resume_path} | CoverLetter: {cover_letter_path}'
    ))
    conn.commit()
    conn.close()
    return True


# ─────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────

def export_to_excel(db_path, outputs_dir):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️ openpyxl not installed — run: pip install openpyxl")
        return

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM applications")
    rows = cursor.fetchall()
    cols = [desc[0] for desc in cursor.description]
    conn.close()

    def make_workbook(data, filename, sheet_title):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title

        headers = [
            "Job Title", "Company", "Salary", "Score",
            "Queue/Reason", "Remote?", "Platform", "Apply Link",
            "Resume Path", "Cover Letter Path", "Date Found", "Status"
        ]
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

        for row in data:
            notes = row[cols.index("notes")] or ""
            url = resume_path = cl_path = queue = ""

            for part in notes.split("|"):
                p = part.strip()
                if p.startswith("URL:"):
                    url = p[4:].strip()
                elif p.startswith("Resume:"):
                    resume_path = p[7:].strip()
                elif p.startswith("CoverLetter:"):
                    cl_path = p[12:].strip()
                elif p.startswith("Queue:"):
                    queue = p[6:].strip()

            jd = row[cols.index("jd_text")] or ""
            is_remote = "✅ Remote" if any(
                w in jd.lower() for w in ["remote", "wfh", "work from home"]
            ) else "On-site/NCR"

            ws.append([
                row[cols.index("job_title")],
                row[cols.index("company")],
                row[cols.index("salary_range")],
                row[cols.index("match_score")],
                queue,
                is_remote,
                row[cols.index("platform")],
                url,
                resume_path,
                cl_path,
                row[cols.index("applied_date")],
                row[cols.index("status")]
            ])

        col_widths = [35, 25, 15, 8, 18, 12, 12, 50, 60, 60, 14, 16]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for row_idx in range(2, ws.max_row + 1):
            url_cell = ws.cell(row=row_idx, column=8)
            if url_cell.value and str(url_cell.value).startswith("http"):
                url_cell.hyperlink = url_cell.value
                url_cell.font = Font(color="0563C1", underline="single")

        path = os.path.join(outputs_dir, filename)
        wb.save(path)
        print(f"  📊 Excel saved: {path}")

    easy_rows   = [r for r in rows if (r[cols.index("status")] or "") == "Discovered"]
    manual_rows = [r for r in rows if (r[cols.index("status")] or "") == "Manual Pending"]

    if easy_rows:
        make_workbook(easy_rows, "easy_apply_jobs.xlsx", "Easy Apply Jobs")
    if manual_rows:
        make_workbook(manual_rows, "manual_apply_jobs.xlsx", "Manual Apply Jobs")

    print(f"  ✅ Excel export done — {len(easy_rows)} easy apply, {len(manual_rows)} manual")


# ─────────────────────────────────────────────
# SCRAPER — PASS 0: EASY APPLY FILTER (Selenium)
# ─────────────────────────────────────────────

def scrape_easy_apply_jobs(driver, max_jobs=25) -> list:
    print("  ⚡ Scraping Easy Apply jobs via f_AL=true filter...")
    jobs = []
    seen_urls = set()
    debug_saved = False  # save debug HTML only once

    for role in TARGET_ROLES:
        for location in EA_SEARCH_LOCATIONS:
            search_query    = role.replace(" ", "%20")
            search_location = location.replace(" ", "%20").replace(",", "%2C")
            url = (
                f"https://www.linkedin.com/jobs/search/"
                f"?keywords={search_query}&location={search_location}"
                f"&f_AL=true&f_TPR=r86400&sortBy=DD"
            )
            print(f"  🔍 Easy Apply: {role} in {location}")

            try:
                driver.get(url)
                time.sleep(random.uniform(3, 4))

                # Wait for job list to render
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR,
                            "ul.jobs-search__results-list, "
                            "div.scaffold-layout__list, "
                            "li.jobs-search-results__list-item"
                        ))
                    )
                except:
                    pass

                for _ in range(3):
                    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(1.5)

                # ── DEBUG: save HTML once to confirm class names ──────
                if not debug_saved:
                    debug_path = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "ea_debug.html"
                    )
                    with open(debug_path, "w", encoding="utf-8") as f:
                        f.write(driver.page_source)
                    print(f"  🔍 Debug HTML saved → ea_debug.html")
                    debug_saved = True
                # ─────────────────────────────────────────────────────

                soup = BeautifulSoup(driver.page_source, "html.parser")

                # Authenticated LinkedIn selectors (left panel job cards)
                job_cards = soup.find_all("li", class_="jobs-search-results__list-item")
                if not job_cards:
                    job_cards = soup.find_all("div", class_="job-card-container")
                if not job_cards:
                    job_cards = soup.find_all("li", attrs={"data-occludable-job-id": True})
                if not job_cards:
                    # fallback — public page selectors
                    job_cards = soup.find_all("div", class_="base-card")
                if not job_cards:
                    job_cards = soup.find_all("li", class_="jobs-search__results-list")

                print(f"  📋 Found {len(job_cards)} Easy Apply cards")

                for card in job_cards[:max_jobs]:
                    try:
                        # Authenticated card selectors
                        title_elem = (
                            card.find("a", class_="job-card-list__title") or
                            card.find("a", class_="job-card-container__link") or
                            card.find("a", attrs={"data-control-name": "job_card_title"})
                        )
                        company_elem = (
                            card.find("span", class_="job-card-container__primary-description") or
                            card.find("span", class_="job-card-container__company-name") or
                            card.find("div",  class_="artdeco-entity-lockup__subtitle")
                        )
                        location_elem = (
                            card.find("li",  class_="job-card-container__metadata-item") or
                            card.find("div", class_="artdeco-entity-lockup__caption") or
                            card.find("span", class_="job-card-container__metadata-item")
                        )
                        salary_elem = card.find("li", class_="job-card-container__metadata-item--salary")
                        date_elem   = card.find("time")

                        if not title_elem:
                            continue

                        title        = title_elem.get_text(strip=True)
                        company      = company_elem.get_text(strip=True)  if company_elem  else "Unknown"
                        job_location = location_elem.get_text(strip=True) if location_elem else location
                        salary_text  = salary_elem.get_text(strip=True)   if salary_elem   else ""
                        posted_date  = date_elem.get("datetime", str(date.today())) if date_elem else str(date.today())

                        href    = title_elem.get("href", "")
                        job_url = (
                            ("https://www.linkedin.com" + href.split("?")[0]).strip("/")
                            if href.startswith("/")
                            else href.split("?")[0].strip("/")
                        )

                        if not job_url or job_url in seen_urls:
                            continue
                        seen_urls.add(job_url)

                        # Extract job_external_id for deduplication
                        job_external_id = extract_job_external_id(job_url)

                        jobs.append({
                            "title":           title,
                            "company":         company,
                            "location":        job_location,
                            "url":             job_url,
                            "job_external_id": job_external_id,
                            "jd_text":         "",
                            "salary_range":    salary_text,
                            "posted_date":     posted_date[:10],
                            "platform":        "linkedin",
                            "easy_apply":      True,
                            "remote":          is_remote_location(job_location),
                            "apply_by_date":  None,  # Will be extracted after JD parsing
                            "expected_salary": estimate_expected_salary(salary_text)
                        })

                    except:
                        continue

            except Exception as e:
                print(f"  ❌ Easy Apply search failed: {e}")
                continue

            time.sleep(random.uniform(2, 3))

    print(f"  ✅ Total Easy Apply jobs found: {len(jobs)}")
    return jobs


# ─────────────────────────────────────────────
# SCRAPER — PASS 1+2: KEYWORD SEARCH (requests)
# ─────────────────────────────────────────────

def scrape_linkedin_jobs(role, location, max_jobs=25):
    jobs = []
    search_query    = role.replace(" ", "%20")
    search_location = location.replace(" ", "%20")
    url = (
        f"https://www.linkedin.com/jobs/search/"
        f"?keywords={search_query}&location={search_location}"
        f"&f_TPR=r86400&sortBy=DD"
    )
    print(f"  🔍 Searching: {role} in {location}")
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        if response.status_code != 200:
            print(f"  ⚠️ LinkedIn returned {response.status_code}")
            return jobs

        soup      = BeautifulSoup(response.text, "html.parser")
        job_cards = soup.find_all("div", class_="base-card")
        if not job_cards:
            job_cards = soup.find_all("li", class_="jobs-search__results-list")

        print(f"  📋 Found {len(job_cards)} cards")

        for card in job_cards[:max_jobs]:
            try:
                title_elem    = card.find("h3", class_="base-search-card__title")
                company_elem  = card.find("h4", class_="base-search-card__subtitle")
                location_elem = card.find("span", class_="job-search-card__location")
                link_elem     = card.find("a", class_="base-card__full-link")
                date_elem     = card.find("time")
                salary_elem   = card.find("span", class_="job-search-card__salary-info")

                title        = title_elem.get_text(strip=True)    if title_elem    else "Unknown"
                company      = company_elem.get_text(strip=True)  if company_elem  else "Unknown"
                job_location = location_elem.get_text(strip=True) if location_elem else location
                job_url      = link_elem.get("href", "").split("?")[0].strip("/") if link_elem else ""
                posted_date  = date_elem.get("datetime", str(date.today())) if date_elem else str(date.today())
                salary_text  = salary_elem.get_text(strip=True) if salary_elem else ""
                
                # Extract job_external_id for deduplication
                job_external_id = extract_job_external_id(job_url)

                jobs.append({
                    "title":           title,
                    "company":         company,
                    "location":        job_location,
                    "url":             job_url,
                    "job_external_id": job_external_id,
                    "jd_text":         "",
                    "salary_range":    salary_text,
                    "posted_date":     posted_date[:10],
                    "platform":        "linkedin",
                    "easy_apply":      False,
                    "remote":          is_remote_location(job_location),
                    "apply_by_date":  None,  # Will be extracted after JD parsing
                    "expected_salary": estimate_expected_salary(salary_text)
                })
            except:
                continue

    except Exception as e:
        print(f"  ❌ Request failed: {e}")

    return jobs


# ─────────────────────────────────────────────
# PER-JOB PROCESSOR
# ─────────────────────────────────────────────

def process_job(driver, job, seen_urls, counters, queues,
                skip_location_filter=False, easy_apply_known=False, total_jobs_processed=0, max_jobs=999999, GENERATE_DOCUMENTS=True):

    # Check job limit
    if total_jobs_processed >= max_jobs:
        return False  # Signal to stop processing
    
    # Progress logging
    print(f"  📝 Processing job {total_jobs_processed + 1}/{max_jobs}: {job['title']} @ {job['company']}")

    # Insert job into database with deduplication
    job_data = {
        "job_title": job["title"],
        "company": job["company"],
        "platform": job["platform"],
        "match_score": None,  # Will be set after scoring
        "job_external_id": job["job_external_id"],
        "easy_apply": 1 if job.get("easy_apply", False) else 0,
        "salary_range": job["salary_range"],
        "status": "Discovered",
        "remote": 1 if job.get("remote", False) else 0,
        "apply_by_date": job.get("apply_by_date"),
        "expected_salary": job.get("expected_salary")
    }
    
    app_id = insert_job(job_data)
    if app_id is None:
        print(f"  🔁 Skipping duplicate job: {job['title']} @ {job['company']} (job_external_id={job['job_external_id']})")
        return True  # Continue processing other jobs

    if not skip_location_filter:
        if not is_ncr_or_remote(job["location"]):
            counters["wfo_skipped"] += 1
            print(f"  🚫 WFO non-NCR: {job['title']} @ {job['company']} ({job['location']})")
            return True

    try:
        driver.get(job["url"])
        time.sleep(random.uniform(2, 3))
    except Exception as e:
        print(f"  ❌ Page load failed: {e}")
        return True

    easy_apply = easy_apply_known or is_easy_apply_page(driver)

    job["jd_text"] = get_jd_text(driver)
    time.sleep(random.uniform(1.0, 2.0))

    # Extract apply_by_date from JD text
    if not job.get("apply_by_date"):
        job["apply_by_date"] = extract_apply_by_date(job["jd_text"])

    salary_lpa = extract_salary_lpa(job["salary_range"])
    dream = is_dream_company(job["company"])
    good_salary = salary_lpa and salary_lpa >= MIN_SALARY_LPA

    # Stage 1: Compute fit_score (non-LLM)
    score_result = score_job(job)
    fit_score = score_result['score']
    print(f"  📊 Fit Score: {fit_score} — {score_result['queue'].upper()} | "
          f"{' | '.join(score_result.get('reasons', []))}")

    # Update database with fit_score and additional metadata
    from database.db import get_connection
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE applications SET match_score = ?, jd_text = ?, apply_by_date = ?, remote = ?, expected_salary = ? WHERE app_id = ?", 
                (fit_score, job["jd_text"], job.get("apply_by_date"), 1 if job.get("remote", False) else 0, job.get("expected_salary"), app_id))
    conn.commit()
    conn.close()

    # Stage 2: Two-stage gating - only generate documents for high-fit jobs if GENERATE_DOCUMENTS is True
    if GENERATE_DOCUMENTS and (fit_score >= AUTO_APPLY_THRESHOLD or dream or good_salary):
        print(f"  ✨ High-fit job (score={fit_score}) – generating ATS-aware resume & cover letter.")
        
        try:
            from resume_tailor import tailor_resume
            from cover_letter import generate_cover_letter
            tailor_result = tailor_resume(
                jd_text=job["jd_text"],
                company_name=job["company"],
                job_title=job["title"]
            )
            resume_path = tailor_result["pdf_path"]
            ats_score = tailor_result["ats_score"]
            
            cl_path = generate_cover_letter(
                jd_text=job["jd_text"],
                company_name=job["company"],
                job_title=job["title"],
                jd_analysis=tailor_result["jd_analysis"]
            )
            
            # Update database with documents and ATS score
            update_job_with_documents(app_id, resume_path=resume_path, 
                                    cover_letter_path=cl_path, ats_score=ats_score)
            
            counters["saved"] += 1
            queue_key = score_result["queue"] if score_result["queue"] in queues else "maybe"
            queues[queue_key].append(
                f'{job["title"]} @ {job["company"]} — {fit_score}/100'
            )
            print(f'  ✅ [EASY/{score_result["queue"].upper()}] '
                  f'{job["title"]} @ {job["company"]} — {fit_score}/100')
            
        except Exception as e:
            print(f"  ❌ Resume/CL generation failed: {e}")
    else:
        if not GENERATE_DOCUMENTS:
            print(f"  🧾 Document generation disabled for this mission (GENERATE_DOCUMENTS=false). Scraping + scoring only.")
        else:
            print(f"  ⏭️ Low-fit job (score={fit_score}) – metadata only, no LLM.")
        counters["no_trigger"] += 1

    return True  # Continue processing


# ─────────────────────────────────────────────
# MAIN RUNNER
# ─────────────────────────────────────────────

def run_job_finder():
    print("🚀 AutoApply mission starting...\n")

    # Determine job limits for mission
    from config import MOCK_MODE
    if MOCK_MODE:
        MAX_JOBS = min(DEV_MAX_JOBS_PER_RUN, MISSION_JOB_BATCH_SIZE)
    else:
        MAX_JOBS = MISSION_JOB_BATCH_SIZE
    
    # Check document generation flag
    GENERATE_DOCUMENTS = os.getenv("GENERATE_DOCUMENTS", str(GENERATE_DOCUMENTS_DEFAULT)).lower() == "true"
    
    total_jobs_processed = 0
    
    print(f"  🎯 Mission: will process up to {MAX_JOBS} new jobs")
    print(f"  📊 Mode: {'Development (MOCK)' if MOCK_MODE else 'Production (Real API)'}")
    print(f"  📄 Document generation: {'Enabled' if GENERATE_DOCUMENTS else 'Disabled (scraping + scoring only)'}")

    counters = {
        "found":         0,
        "wfo_skipped":   0,
        "no_trigger":    0,
        "saved":         0,
        "dream_manual":  0,
        "salary_manual": 0,
    }
    queues    = {"dream": [], "auto_apply": [], "maybe": []}
    seen_urls = set()

    print("  🌐 Opening Chrome (dedicated automation profile)...")
    driver = create_driver()
    print("  ✅ Chrome ready!\n")

    try:
        login_linkedin(driver)

        # ── PASS 0: Easy Apply filter (Selenium, logged in) ───────────
        print("\n⚡ PASS 0 — Easy Apply Jobs (f_AL=true)\n")
        ea_jobs = scrape_easy_apply_jobs(driver, max_jobs=25)
        time.sleep(random.uniform(2, 3))

        for job in ea_jobs:
            if total_jobs_processed >= MAX_JOBS:
                break
                
            should_continue = process_job(
                driver, job, seen_urls, counters, queues,
                skip_location_filter=("remote" in job["location"].lower()),
                easy_apply_known=True,
                total_jobs_processed=total_jobs_processed,
                max_jobs=MAX_JOBS,
                GENERATE_DOCUMENTS=GENERATE_DOCUMENTS
            )
            if should_continue:
                total_jobs_processed += 1
            else:
                break  # Hit job limit

        # ── PASS 1: NCR keyword searches (requests, dream/salary) ─────
        if total_jobs_processed < MAX_JOBS:
            print("\n📍 PASS 1 — NCR Searches (Dream/Salary)\n")
            for role in TARGET_ROLES:
                for location in NCR_SEARCH_LOCATIONS:
                    if total_jobs_processed >= MAX_JOBS:
                        break
                        
                    jobs = scrape_linkedin_jobs(role, location, max_jobs=25)
                    time.sleep(random.uniform(2, 4))

                    eligible = [j for j in jobs if is_ncr_or_remote(j["location"])]
                    batch_skipped = len(jobs) - len(eligible)
                    if batch_skipped:
                        counters["wfo_skipped"] += batch_skipped
                        print(f"  🚫 Batch filtered {batch_skipped} WFO non-NCR instantly")

                    for job in eligible:
                        if total_jobs_processed >= MAX_JOBS:
                            break
                            
                        should_continue = process_job(driver, job, seen_urls, counters, queues,
                                                    skip_location_filter=False,
                                                    total_jobs_processed=total_jobs_processed,
                                                    max_jobs=MAX_JOBS,
                                                    GENERATE_DOCUMENTS=GENERATE_DOCUMENTS)
                        if should_continue:
                            total_jobs_processed += 1
                        else:
                            break  # Hit job limit
                    
                    if total_jobs_processed >= MAX_JOBS:
                        break

        # ── PASS 2: Remote/Worldwide (requests, dream/salary) ─────────
        if total_jobs_processed < MAX_JOBS:
            print("\n🌐 PASS 2 — Remote / Worldwide Searches\n")
            for role in TARGET_ROLES:
                for location in REMOTE_SEARCH_LOCATIONS:
                    if total_jobs_processed >= MAX_JOBS:
                        break
                        
                    jobs = scrape_linkedin_jobs(role, location, max_jobs=25)
                    time.sleep(random.uniform(2, 4))

                    for job in jobs:
                        if total_jobs_processed >= MAX_JOBS:
                            break
                            
                        should_continue = process_job(driver, job, seen_urls, counters, queues,
                                                    skip_location_filter=True,
                                                    total_jobs_processed=total_jobs_processed,
                                                    max_jobs=MAX_JOBS,
                                                    GENERATE_DOCUMENTS=GENERATE_DOCUMENTS)
                        if should_continue:
                            total_jobs_processed += 1
                        else:
                            break  # Hit job limit
                    
                    if total_jobs_processed >= MAX_JOBS:
                        break

    finally:
        driver.quit()
        print("\n  🌐 Browser closed.")

    # ── Excel Export ──────────────────────────────────────────────────
    print("\n📊 Exporting to Excel...")
    export_to_excel(DB_PATH, OUTPUTS_DIR)

    # ── Summary ───────────────────────────────────────────────────────
    print(f"\n{'─'*55}")
    print(f"🎯 Mission Complete")
    print(f"{'─'*55}")
    print(f"  Jobs processed   : {total_jobs_processed}/{MAX_JOBS}")
    print(f"  Total found       : {counters['found']}")
    print(f"  WFO non-NCR skip  : {counters['wfo_skipped']}")
    print(f"  No trigger skip   : {counters['no_trigger']}")
    print(f"  Saved (Easy Apply): {counters['saved']}")
    print(f"  Dream manual pkg  : {counters['dream_manual']}")
    print(f"  Salary manual pkg : {counters['salary_manual']}")
    print(f"{'─'*55}")
    print(f"  ⚡ Easy Apply: {len(queues['auto_apply']) + len(queues['maybe']) + len(queues['dream'])}")
    print(f"  🌟 Dream     : {len(queues['dream'])}")
    print(f"  ✅ Auto-Apply: {len(queues['auto_apply'])}")
    print(f"  👀 Maybe     : {len(queues['maybe'])}")
    print(f"{'─'*55}")
    
    # Mission complete message
    if total_jobs_processed >= MAX_JOBS:
        print(f"✅ Mission complete: processed {total_jobs_processed} new jobs.")
    else:
        print(f"✅ Mission complete: processed {total_jobs_processed} new jobs (no more jobs found).")
    
    print(f"\n📋 To fetch {MAX_JOBS} more new jobs, run this script again")
    print(f"{'─'*55}\n")

    return queues


if __name__ == "__main__":
    run_job_finder()
